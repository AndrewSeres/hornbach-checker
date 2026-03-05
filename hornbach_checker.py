"""
Hornbach Brikety – Kontrola dostupnosti v predajniach
=====================================================
Požiadavky: pip install playwright openpyxl
            python -m playwright install chromium

Spustenie: python hornbach_checker.py
"""

import asyncio
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import tkinter as tk
    from tkinter import ttk, scrolledtext, messagebox
    HAS_GUI = True
except ImportError:
    HAS_GUI = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Chyba openpyxl. Spusti: pip install openpyxl")
    sys.exit(1)

try:
    from playwright.async_api import async_playwright
except ImportError:
    print("Chyba playwright. Spusti: pip install playwright && python -m playwright install chromium")
    sys.exit(1)

# ── Konfigurácia ──────────────────────────────────────────────────────────────
CATEGORY_URL = "https://www.hornbach.sk/c/krby-radiatory-a-klimatizacie/palivove-drevo-pelety-a-brikety/S15929/"
KEYWORDS = ["briketa", "brikety", "brikiet", "briket"]
# Exclude hnedouhoľné
EXCLUDE_KEYWORDS = ["hnedouholn", "uholn", "hnedouhol"]

# Kanonické názvy predajní – slúžia na dedup aj zoradenie
CANONICAL_STORES = [
    "HORNBACH Nitra",
    "HORNBACH Bratislava - Ružinov",
    "HORNBACH Bratislava - Devínska Nová Ves",
    "HORNBACH Košice",
    "HORNBACH Prešov",
]


def matches_keywords(text: str) -> bool:
    t = text.lower()
    t_plain = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode("ascii")
    # Must match at least one keyword
    if not any(k in t or k in t_plain for k in KEYWORDS):
        return False
    # Must NOT match any exclude keyword
    if any(k in t or k in t_plain for k in EXCLUDE_KEYWORDS):
        return False
    return True


def canonicalize_store(raw_name: str):
    """Mapuje raw názov predajne z modalu na kanonický názov.
    Vracia None ak to nie je predajňa (napr. 'Všetky predajne...').
    """
    raw_lower = raw_name.lower().strip()

    # Filtruj "Všetky predajne" a podobné
    if "etky predajne" in raw_lower or "vsetky predajne" in raw_lower:
        return None

    # Filtruj ak neobsahuje HORNBACH
    if "hornbach" not in raw_lower:
        return None

    # Matchni voči kanonickým názvom
    for canonical in CANONICAL_STORES:
        canon_lower = canonical.lower()
        if raw_lower.startswith(canon_lower):
            return canonical
        # Fallback: posledné kľúčové slovo z kanonického názvu
        city_part = canon_lower.replace("hornbach ", "").strip()
        city_keywords = [w.strip() for w in city_part.split("-")]
        last_keyword = city_keywords[-1].strip()
        if last_keyword and last_keyword in raw_lower:
            return canonical

    # Neznáma predajňa – použi meno po prvú čiarku
    base = raw_name.split(",")[0].strip().rstrip(".")
    return base


def normalize_store_display(name: str) -> str:
    """Krátky display názov pre Excel."""
    n = name.replace("HORNBACH", "").strip()
    n = n.lstrip("- ").strip()
    return re.sub(r'\s+', ' ', n).strip() or name


def store_sort_key(s):
    for i, canonical in enumerate(CANONICAL_STORES):
        if s == canonical:
            return (i, s)
    return (99, s)


# ── Scraper ───────────────────────────────────────────────────────────────────
async def scrape(log_fn=print, progress_fn=None):
    products = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
            locale="sk-SK",
        )
        page = await context.new_page()

        # ── 1. Kategória ──────────────────────────────────────────────────────
        log_fn("Nacitavam kategoriu brikety...")
        await page.goto(CATEGORY_URL, wait_until="networkidle", timeout=60000)

        # Cookie banner
        for sel in [
            "#onetrust-accept-btn-handler",
            "button:has-text('Prijat')", "button:has-text('Suhlasim')",
            "button:has-text('Akceptovat')", "button:has-text('Prijať')",
            "button:has-text('Súhlasím')",
        ]:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible(timeout=2000):
                    await btn.click()
                    await page.wait_for_timeout(800)
                    break
            except Exception:
                pass

        # ── 2. Scroll + zbieraj URL ───────────────────────────────────────────
        log_fn("Scrollujem stranku a zbieram produkty...")

        for _ in range(8):
            await page.evaluate("window.scrollBy(0, 800)")
            await page.wait_for_timeout(600)
        await page.evaluate("window.scrollTo(0, 0)")
        await page.wait_for_timeout(500)

        seen_urls = set()
        product_links = []

        all_links = await page.query_selector_all("a[href*='/p/']")
        log_fn(f"  Celkovo linkov na stranke: {len(all_links)}")

        for link in all_links:
            try:
                href = await link.get_attribute("href")
                if not href or '/p/' not in href:
                    continue
                full_url = href if href.startswith("http") else f"https://www.hornbach.sk{href}"
                if full_url in seen_urls:
                    continue

                text = ""
                try:
                    text = (await link.inner_text()).strip()
                except Exception:
                    pass

                if not text or len(text) < 5:
                    try:
                        text = await page.evaluate("""el => {
                            const tile = el.closest('article') ||
                                         el.closest('li') ||
                                         el.closest('[class*="Tile"]') ||
                                         el.closest('[class*="tile"]') ||
                                         el.closest('[class*="product"]') ||
                                         el.parentElement;
                            return tile ? tile.innerText : '';
                        }""", link)
                        text = (text or "").strip()
                    except Exception:
                        pass

                if not text:
                    continue

                if matches_keywords(text):
                    seen_urls.add(full_url)
                    name = next(
                        (l.strip() for l in text.splitlines() if len(l.strip()) > 8),
                        text[:80]
                    )
                    product_links.append({"url": full_url, "name": name[:100]})
                    log_fn(f"  + {name[:65]}")
            except Exception:
                pass

        log_fn(f"\nCelkom {len(product_links)} produktov na spracovanie")

        if not product_links:
            log_fn("Ziadne produkty nenajdene. Skontroluj URL alebo klucove slova.")
            await browser.close()
            return products

        # ── 3. Detail produktu ────────────────────────────────────────────────
        total = len(product_links)
        for idx, prod in enumerate(product_links):
            pct = int((idx / total) * 100)
            if progress_fn:
                progress_fn(pct, f"Produkt {idx+1}/{total}: {prod['name'][:45]}")
            log_fn(f"\n[{idx+1}/{total}] {prod['name']}")

            prod_data = {
                "name": prod["name"],
                "ean": "",
                "artikel_nr": "",
                "price": "",
                "url": prod["url"],
                "stores": {}
            }

            m = re.search(r'/(\d{5,})/?', prod["url"])
            if m:
                prod_data["artikel_nr"] = m.group(1)

            pp = None
            try:
                pp = await context.new_page()
                await pp.goto(prod["url"], wait_until="domcontentloaded", timeout=45000)
                await pp.wait_for_timeout(2000)

                # Cena
                try:
                    for price_sel in [
                        "[data-testid='article-price'] [class*='value']",
                        "[class*='article-price'] [class*='value']",
                        "[class*='ArticlePrice'] [class*='value']",
                        "[class*='price-value']",
                        "[class*='Price'] strong",
                        "[class*='price'] strong",
                        "strong[class*='price']",
                    ]:
                        el = pp.locator(price_sel).first
                        if await el.is_visible(timeout=800):
                            prod_data["price"] = (await el.inner_text()).strip()
                            break
                except Exception:
                    pass

                # Cena fallback: hľadaj regex v texte stránky
                if not prod_data["price"]:
                    try:
                        body_text = await pp.inner_text("body")
                        # Match "4,59 €" or "4,59 €*/ks" etc.
                        pm = re.search(r'(\d+,\d{2})\s*€', body_text)
                        if pm:
                            prod_data["price"] = f"{pm.group(1)} €"
                    except Exception:
                        pass

                # Rozbalit EAN sekciu
                for expand_text in [
                    "VIAC INFORMACII O VYROBKU", "VIAC INFORMÁCIÍ O VÝROBKU",
                    "Viac informacii", "Viac informácií",
                    "ZOBRAZIT VIAC", "ZOBRAZIŤ VIAC",
                    "Zobrazit viac", "Zobraziť viac",
                ]:
                    try:
                        btn = pp.get_by_text(expand_text, exact=False).first
                        if await btn.is_visible(timeout=800):
                            await btn.click()
                            await pp.wait_for_timeout(700)
                            break
                    except Exception:
                        pass

                # EAN z tabulky
                try:
                    rows = await pp.query_selector_all("tr")
                    for row in rows:
                        row_text = await row.inner_text()
                        if "EAN" in row_text:
                            cells = await row.query_selector_all("td")
                            if len(cells) >= 2:
                                prod_data["ean"] = (await cells[-1].inner_text()).strip()
                            elif len(cells) == 1:
                                em = re.search(r'EAN\s*[:\s]+([0-9, ]+)', row_text)
                                if em:
                                    prod_data["ean"] = em.group(1).strip()
                            if prod_data["ean"]:
                                break
                except Exception:
                    pass

                # EAN fallback
                if not prod_data["ean"]:
                    try:
                        body = await pp.inner_text("body")
                        em2 = re.search(r'EAN\s*\n?\s*([0-9]{8,}(?:[,\s]+[0-9]{8,})*)', body)
                        if em2:
                            prod_data["ean"] = em2.group(1).strip()
                    except Exception:
                        pass

                log_fn(f"  EAN: {prod_data['ean'] or '-'}  Cena: {prod_data['price'] or '-'}")

                # Klikni na dostupnost
                clicked = False
                for avail_text in [
                    "SKONTROLOVAT DOSTUPNOST V NAJBLIZEJ PREDAJNI",
                    "SKONTROLOVAT DOSTUPNOST",
                    "SKONTROLOVAŤ DOSTUPNOSŤ V NAJBLIŽŠEJ PREDAJNI",
                    "SKONTROLOVAŤ DOSTUPNOSŤ",
                    "Skontrolovat dostupnost",
                    "Skontrolovať dostupnosť",
                    "dostupnost v predajni",
                    "dostupnosť v predajni",
                ]:
                    try:
                        btn = pp.get_by_text(avail_text, exact=False).first
                        if await btn.is_visible(timeout=1500):
                            await btn.click()
                            await pp.wait_for_timeout(3000)
                            clicked = True
                            log_fn("  Otvoreny modal dostupnosti")
                            break
                    except Exception:
                        pass

                if not clicked:
                    for btn_sel in [
                        "[class*='StoreAvailability'] button",
                        "[class*='store-availability'] button",
                        "button[class*='store']",
                    ]:
                        try:
                            btn = pp.locator(btn_sel).first
                            if await btn.is_visible(timeout=800):
                                await btn.click()
                                await pp.wait_for_timeout(3000)
                                clicked = True
                                break
                        except Exception:
                            pass

                # Citaj modal
                modal_text = ""
                for modal_sel in ["[role='dialog']", "[class*='Modal']", "[class*='modal']", "[class*='Overlay']"]:
                    try:
                        modal = pp.locator(modal_sel).first
                        if await modal.is_visible(timeout=2000):
                            t = await modal.inner_text()
                            if "HORNBACH" in t:
                                modal_text = t
                                break
                    except Exception:
                        pass

                # Fallback: cela stranka
                if not modal_text or "HORNBACH" not in modal_text:
                    try:
                        modal_text = await pp.inner_text("body")
                    except Exception:
                        pass

                # ── Parsuj predajne (FIX: kanonizácia + lepší stock parsing) ──
                if modal_text and "HORNBACH" in modal_text:
                    lines = [l.strip() for l in modal_text.splitlines() if l.strip()]
                    found_stores = {}  # kanonický_názov -> stock
                    i = 0
                    while i < len(lines):
                        line = lines[i]
                        if "HORNBACH" in line:
                            canonical = canonicalize_store(line)
                            if canonical is not None:
                                # Hľadaj stock v nasledujúcich riadkoch
                                stock_val = None
                                for j in range(i + 1, min(i + 15, len(lines))):
                                    jline = lines[j]

                                    # Stop ak narazíme na ďalšiu predajňu
                                    if "HORNBACH" in jline and j > i:
                                        break

                                    # "123 balenie/balení/ks/kusov"
                                    sm = re.search(
                                        r'(\d[\d\s]*)\s*(balen[ií]e?|ks\b|kus)',
                                        jline, re.IGNORECASE
                                    )
                                    if sm:
                                        stock_val = int(sm.group(1).replace(" ", ""))
                                        break

                                    # Samostatné číslo na riadku + kontext
                                    if re.match(r'^\d+$', jline):
                                        nearby = " ".join(lines[max(0, j-2):min(len(lines), j+3)]).lower()
                                        if any(w in nearby for w in [
                                            "dostupn", "skladom", "dispoz",
                                            "baleni", "balení", "objedn"
                                        ]):
                                            stock_val = int(jline)
                                            break

                                    # Nedostupné
                                    if re.search(
                                        r'nie je k dispoz|nedostupn|momentálne nie|'
                                        r'momentalne nie|v súčasnosti|v sucasnosti|'
                                        r'vypredané|vypredane|0 balení|0 balenie',
                                        jline, re.IGNORECASE
                                    ):
                                        stock_val = 0
                                        break

                                    # "Dostupné > 50 balenie" alebo "Dostupné 1 balenie"
                                    sm2 = re.search(
                                        r'dostupn[éeý]\s*>?\s*(\d+)',
                                        jline, re.IGNORECASE
                                    )
                                    if sm2:
                                        stock_val = int(sm2.group(1))
                                        break

                                # Ak sme už videli túto predajňu
                                if canonical in found_stores:
                                    # Updatni len ak predchadzajúci bol "?" a teraz máme číslo
                                    if found_stores[canonical] == "?" and stock_val is not None:
                                        found_stores[canonical] = stock_val
                                        log_fn(f"  {canonical} (update): {stock_val}")
                                else:
                                    if stock_val is None:
                                        stock_val = "?"
                                    found_stores[canonical] = stock_val
                                    log_fn(f"  {canonical}: {stock_val}")
                        i += 1

                    prod_data["stores"] = found_stores

                if not prod_data["stores"]:
                    log_fn("  Predajne nenajdene")

            except Exception as e:
                log_fn(f"  Chyba: {e}")
            finally:
                if pp:
                    try:
                        await pp.close()
                    except Exception:
                        pass

            products.append(prod_data)
            await asyncio.sleep(0.5)

        await browser.close()

    if progress_fn:
        progress_fn(100, "Hotovo!")
    log_fn(f"\nHotovo – {len(products)} produktov spracovanych")
    return products


# ── Excel export (pivot: 1 riadok = 1 produkt, predajne ako stĺpce) ──────────
def export_excel(products: list) -> str:
    if not products:
        return ""

    # Zozbieraj a zoraď predajne
    all_stores_set = set()
    for p in products:
        all_stores_set.update(p["stores"].keys())
    all_stores = sorted(all_stores_set, key=store_sort_key)

    wb = Workbook()

    ORANGE    = "F4600C"
    DARK_OR   = "C44D00"
    STORE_HDR = "7B2D00"
    ROW_A     = "FFFFFF"
    ROW_B     = "FFF3EE"
    GREEN_BG  = "C6EFCE"
    YELLOW_BG = "FFEB9C"
    RED_BG    = "FFC7CE"
    GREY_BG   = "EEEEEE"
    GREEN_FG  = "276221"
    YELLOW_FG = "7D6608"
    RED_FG    = "9C0006"

    thin  = Side(style="thin",   color="DDDDDD")
    thick = Side(style="medium", color=DARK_OR)

    # ── List 1: Prehľad dostupnosti ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Prehľad dostupnosti"

    fixed = ["Produkt", "EAN", "Artikl č.", "Cena"]

    # Hlavičky – fixné stĺpce
    for ci, title in enumerate(fixed, 1):
        c = ws.cell(row=1, column=ci, value=title)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(bottom=thick, right=Side(style="thin", color="CC4400"))

    # Hlavičky – predajne
    for si, store in enumerate(all_stores):
        c = ws.cell(row=1, column=len(fixed) + 1 + si, value=normalize_store_display(store))
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor=STORE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(bottom=thick, right=Side(style="thin", color="AA3300"))

    # Hlavička – URL (posledný stĺpec)
    url_col = len(fixed) + len(all_stores) + 1
    c = ws.cell(row=1, column=url_col, value="URL")
    c.font      = Font(bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor=ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = Border(bottom=thick, right=Side(style="thin", color="CC4400"))

    ws.row_dimensions[1].height = 34

    # Dáta
    for ri, prod in enumerate(products, 2):
        bg = ROW_A if ri % 2 == 0 else ROW_B

        def dc(col, val, bold=False, center=False, wrap=False):
            c = ws.cell(row=ri, column=col, value=val)
            c.font      = Font(size=10, bold=bold)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if center else "left",
                                    wrap_text=wrap)
            c.border    = Border(right=thin, bottom=thin)
            return c

        dc(1, prod["name"], wrap=True)
        dc(2, prod["ean"])
        dc(3, prod["artikel_nr"], center=True)
        dc(4, prod["price"], center=True)

        for si, store in enumerate(all_stores):
            col   = len(fixed) + 1 + si
            stock = prod["stores"].get(store, "–")
            c     = ws.cell(row=ri, column=col, value=stock)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = Border(right=thin, bottom=thin)
            try:
                n = int(stock)
                if n == 0:
                    c.fill = PatternFill("solid", fgColor=RED_BG)
                    c.font = Font(size=10, bold=True, color=RED_FG)
                elif n < 20:
                    c.fill = PatternFill("solid", fgColor=YELLOW_BG)
                    c.font = Font(size=10, bold=True, color=YELLOW_FG)
                else:
                    c.fill = PatternFill("solid", fgColor=GREEN_BG)
                    c.font = Font(size=10, bold=True, color=GREEN_FG)
            except (ValueError, TypeError):
                c.fill = PatternFill("solid", fgColor=GREY_BG)
                c.font = Font(size=10, color="888888")

        dc(url_col, prod["url"])
        ws.row_dimensions[ri].height = 22

    # Šírky stĺpcov
    for ci, w in enumerate([42, 22, 10, 10], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for si in range(len(all_stores)):
        ws.column_dimensions[get_column_letter(len(fixed) + 1 + si)].width = 16
    ws.column_dimensions[get_column_letter(url_col)].width = 45

    # Freeze: prvý riadok + fixné stĺpce
    ws.freeze_panes = f"{get_column_letter(len(fixed) + 1)}2"

    # ── List 2: Súhrn ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Súhrn")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22

    ws2.cell(row=1, column=1, value="Súhrn – Hornbach Brikety").font = Font(bold=True, size=13, color=ORANGE)
    ws2.row_dimensions[1].height = 26

    date_str_display = datetime.now().strftime("%d.%m.%Y %H:%M")

    for r, (lbl, val) in enumerate([
        ("Dátum kontroly", date_str_display),
        ("Počet produktov", len(products)),
        ("Počet predajní",  len(all_stores)),
    ], 3):
        ws2.cell(row=r, column=1, value=lbl).font = Font(size=11)
        c = ws2.cell(row=r, column=2, value=val)
        c.font = Font(size=11, bold=True)

    ws2.cell(row=7, column=1, value="Predajňa").font = Font(bold=True, size=11)
    ws2.cell(row=7, column=2, value="Celkové zásoby").font = Font(bold=True, size=11)
    for ri2, store in enumerate(all_stores, 8):
        total_stock = sum(
            int(p["stores"].get(store, 0)) for p in products
            if isinstance(p["stores"].get(store), int)
        )
        ws2.cell(row=ri2, column=1, value=normalize_store_display(store)).font = Font(size=10)
        c = ws2.cell(row=ri2, column=2, value=total_stock)
        c.font = Font(size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=(GREEN_BG if total_stock > 0 else RED_BG))
        c.alignment = Alignment(horizontal="center")

    # Ulož
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"hornbach_brikety_{date_str}.xlsx"
    desktop  = Path.home() / "Desktop"
    out      = (desktop / filename) if desktop.exists() else (Path.home() / filename)
    wb.save(str(out))
    return str(out)


# ── GUI ───────────────────────────────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        root.title("Hornbach Brikety Checker")
        root.geometry("720x580")
        root.configure(bg="#1a1a1a")
        root.resizable(True, True)

        hdr = tk.Frame(root, bg="#F4600C", height=56)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🔥  Hornbach Brikety – Kontrola skladu",
                 bg="#F4600C", fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=20, pady=14)

        pb_f = tk.Frame(root, bg="#1a1a1a", pady=10)
        pb_f.pack(fill="x", padx=20)
        self.pb_label = tk.Label(pb_f, text="Pripraveny", bg="#1a1a1a", fg="#888",
                                  font=("Consolas", 9))
        self.pb_label.pack(anchor="w")
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TProgressbar", troughcolor="#2a2a2a", background="#F4600C", thickness=7)
        self.pb = ttk.Progressbar(pb_f, length=680, mode="determinate")
        self.pb.pack(fill="x", pady=(3, 0))

        log_f = tk.Frame(root, bg="#1a1a1a")
        log_f.pack(fill="both", expand=True, padx=20)
        tk.Label(log_f, text="Log:", bg="#1a1a1a", fg="#555",
                 font=("Consolas", 8)).pack(anchor="w")
        self.log = scrolledtext.ScrolledText(
            log_f, bg="#0d0d0d", fg="#d4d4d4",
            font=("Consolas", 9), insertbackground="white",
            relief="flat", borderwidth=1,
            highlightbackground="#333", state="disabled", height=18)
        self.log.pack(fill="both", expand=True)

        btn_f = tk.Frame(root, bg="#1a1a1a", pady=12)
        btn_f.pack(fill="x", padx=20)

        self.btn_run = tk.Button(
            btn_f, text="▶  SPUSTIT KONTROLU", command=self.start,
            bg="#F4600C", fg="white", font=("Segoe UI", 11, "bold"),
            relief="flat", padx=22, pady=9, cursor="hand2",
            activebackground="#C44D00", activeforeground="white")
        self.btn_run.pack(side="left")

        self.btn_export = tk.Button(
            btn_f, text="↓  Exportovat Excel", command=self.do_export,
            bg="#2a2a2a", fg="#888", font=("Segoe UI", 10),
            relief="flat", padx=16, pady=9, cursor="hand2", state="disabled")
        self.btn_export.pack(side="left", padx=(10, 0))

        self.status = tk.Label(btn_f, text="", bg="#1a1a1a", fg="#888",
                                font=("Consolas", 9))
        self.status.pack(side="right")
        self.products = []

    def log_msg(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")
        self.root.update_idletasks()

    def set_progress(self, pct, label=""):
        self.pb["value"] = pct
        self.pb_label.configure(text=label or f"{pct}%")
        self.root.update_idletasks()

    def start(self):
        self.btn_run.configure(state="disabled", text="⏳ Prebieha...")
        self.btn_export.configure(state="disabled")
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")
        self.products = []

        import threading
        def run():
            try:
                self.products = asyncio.run(
                    scrape(log_fn=self.log_msg, progress_fn=self.set_progress))
                self.root.after(0, self.on_done)
            except Exception as e:
                self.log_msg(f"\nKriticka chyba: {e}")
                self.root.after(0, lambda: self.btn_run.configure(
                    state="normal", text="▶  SPUSTIT KONTROLU"))

        threading.Thread(target=run, daemon=True).start()

    def on_done(self):
        self.btn_run.configure(state="normal", text="▶  SPUSTIT KONTROLU")
        if self.products:
            self.btn_export.configure(state="normal", bg="#F4600C", fg="white")
            self.status.configure(text=f"✓ {len(self.products)} produktov", fg="#4caf50")
            messagebox.showinfo("Hotovo",
                f"Spracovanych {len(self.products)} produktov.\n"
                "Klikni '↓ Exportovat Excel'.")
        else:
            self.status.configure(text="Ziadne vysledky", fg="#ffc107")

    def do_export(self):
        if not self.products:
            messagebox.showwarning("Prazdne", "Najprv spusti kontrolu.")
            return
        try:
            path = export_excel(self.products)
            self.log_msg(f"\nExcel ulozeny: {path}")
            messagebox.showinfo("Excel ulozeny", f"Subor ulozeny:\n{path}")
        except Exception as e:
            messagebox.showerror("Chyba exportu", str(e))


if __name__ == "__main__":
    if HAS_GUI:
        root = tk.Tk()
        app = App(root)
        root.mainloop()
    else:
        results = asyncio.run(scrape())
        if results:
            path = export_excel(results)
            print(f"Excel ulozeny: {path}")
